from openpyxl import Workbook
from openpyxl import load_workbook

# charger le classeur 'sample' dans variable wb
wb = load_workbook("sample.xlsx")

# ouvrir la feuille 'Feuil1' dans variable wb
ws = wb["Feuil1"]

# accès à la valeur d'une cellule
a = ws["A2"].value
print(a)

# accès à la valeur d'une cellule
b = ws["B2"].value
print(b)

# ecrire une valeur dans une cellule C2
print(a*b)
ws["C2"].value = a * b

# enregistrer les modificiation dans le classseur 'sampleNew'
wb.save("sampleAfter.xlsx")

print("Merci. Ouvrez SampleAfter.xlsx")


# Télécharger le fichier sampleAfter.xlsx
# Par GitHub
# Source :
# https://www.tutsmake.com/upload-project-files-on-github-using-command-line/

# 1- Sur rnavigateur, aller sur https://github.com/
#    Créer un compte
#       -> Retenir le nom de l'utilisateur et son mail ici :
#          VulgaireBidon
#          vulgaire.bidon@gmail.com

#    Créer un repository
#       -> Retenir l'URL du repository ici : 
#          https://github.com/VulgaireBidon/pythonL3EEM.git

# 2- Dans l'onglet SHELL de ReplIt
#    taper (sans le #) les 8 commandes suivantes puis presser ENTREE à chaque ligne
#    en adaptant les 3 informations personnelles à son propre cas
#    1)      VulgaireBidon
#    2)      vulgaire.bidon@gmail.com
#    3)      https://github.com/VulgaireBidon/pythonL3EEM.git
#


# git init
# git add .
# git config user.email "vulgaire.bidon@gmail.com"
# git config user.name "VulgaireBidon"
# git commit -m "Fichier Projet"
# git branch -M master
# git remote add origin https://github.com/VulgaireBidon/pythonL3EEM.git
# git push -u origin master


# 3- Sur navigateur Visiter alors la page du repository     https://github.com/VulgaireBidon/pythonL3EEM.git

# 4- les fichiers sont téléchageables